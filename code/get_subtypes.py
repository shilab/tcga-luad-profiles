from __future__ import print_function
from openpyxl import load_workbook
import warnings

warnings.simplefilter("ignore")
wb = load_workbook(filename='data/nature13385-s2.xlsx', read_only=True)
ws = wb['S_Table 7-Clinical&Molec_Summar']

ids = []
subtype = []

for row in ws['A6':'A235']:
    for cell in row:
        ids.append(cell.value)

for row in ws['AK6':'AK235']:
    for cell in row:
        subtype.append(cell.value)

for counter, id in enumerate(ids):
    print(id + "\t" + subtype[counter])
